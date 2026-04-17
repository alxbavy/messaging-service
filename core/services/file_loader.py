from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from core.domain.dto import ClientContactDTO
from core.domain.enums import ImportFileHeader
from core.domain.normalizers import (
    clean_string,
    normalize_email,
    normalize_handle,
    normalize_phone,
    normalize_telegram,
)
from core.models.client import Client, ClientContact
from core.models.enums import ContactLabel, ContactStatus, ContactType


@dataclass(slots=True)
class ImportStats:
    created: int = 0
    updated: int = 0
    contacts_created: int = 0
    contacts_updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


class FileLoaderService:
    SUPPORTED_SUFFIXES = {".xlsx"}

    def import_file(self, file_path: str | Path) -> ImportStats:
        path = Path(file_path)
        self._validate_path(path)

        stats = ImportStats()
        workbook = load_workbook(filename=path, read_only=True, data_only=True)

        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)

            try:
                headers = next(rows)
            except StopIteration as exc:
                raise ValueError("Файл пустой") from exc

            header_map = self._build_header_map(headers)

            for row_index, row in enumerate(rows, start=2):
                if self._is_empty_row(row):
                    continue

                try:
                    self._import_row(row=row, header_map=header_map, stats=stats)
                except Exception as exc:
                    stats.skipped += 1
                    stats.errors.append(f"Строка {row_index}: {exc}")
        finally:
            workbook.close()

        return stats

    def _validate_path(self, path: Path) -> None:
        if not path.exists():
            raise FileNotFoundError(f"Файл не найден: {path}")

        if not path.is_file():
            raise ValueError(f"Ожидался файл, получен путь: {path}")

        if path.suffix.lower() not in self.SUPPORTED_SUFFIXES:
            raise ValueError(f"Неподдерживаемый формат файла: {path.suffix}")

    def _build_header_map(self, headers: tuple[Any, ...]) -> dict[str, int]:
        result: dict[str, int] = {}

        for index, header in enumerate(headers):
            header_name = clean_string(header)
            if not header_name:
                continue
            result[header_name] = index

        return result

    def _is_empty_row(self, row: tuple[Any, ...]) -> bool:
        return all(clean_string(value) is None for value in row)

    def _cell(self, row: tuple[Any, ...], header_map: dict[str, int], name: str) -> Any:
        index = header_map.get(name)
        if index is None:
            raise ValueError('Неизвестный заголовок колонки')
        if index >= len(row):
            return None
        return row[index]

    def _import_row(
        self,
        *,
        row: tuple[Any, ...],
        header_map: dict[str, int],
        stats: ImportStats,
    ) -> None:
        external_id = self._parse_external_id(
            self._cell(row, header_map, ImportFileHeader.CLIENT_ID)
        )

        if external_id is None:
            raise ValueError(f'Поле \"{ImportFileHeader.CLIENT_ID}\" должно быть заполнено')

        client_defaults = {
            "full_name": clean_string(
                self._cell(row, header_map, ImportFileHeader.FULL_NAME)
            ),
            "birth_date": self._parse_date(
                self._cell(row, header_map, ImportFileHeader.BIRTH_DATE)
            ),
            "gender": clean_string(
                self._cell(row, header_map, ImportFileHeader.GENDER)
            ) or "",
            "parent_primary": clean_string(
                self._cell(row, header_map, ImportFileHeader.PARENT_PRIMARY)
            ) or "",
            "parent_secondary": clean_string(
                self._cell(row, header_map, ImportFileHeader.PARENT_SECONDARY)
            ) or "",
            "student_status": clean_string(
                self._cell(row, header_map, ImportFileHeader.STATUS)
            ) or "",
            "student_tag": clean_string(
                self._cell(row, header_map, ImportFileHeader.TAG)
            ) or "",
            "student_branch": clean_string(
                self._cell(row, header_map, ImportFileHeader.BRANCH)
            ) or "",
            "can_receive": self._parse_bool(
                self._cell(row, header_map, ImportFileHeader.CAN_RECEIVE)
            ),
        }

        contacts: list[ClientContactDTO] = []

        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.PHONE,
            label="",
            value=normalize_phone(
                self._cell(row, header_map, ImportFileHeader.PHONE_CHILD)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.PHONE,
            label=ContactLabel.SECONDARY,
            value=normalize_phone(
                self._cell(row, header_map, ImportFileHeader.PHONE_SECONDARY)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.PHONE,
            label=ContactLabel.EXTRA,
            value=normalize_phone(
                self._cell(row, header_map, ImportFileHeader.PHONE_EXTRA)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.EMAIL,
            label="",
            value=normalize_email(
                self._cell(row, header_map, ImportFileHeader.EMAIL)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.TELEGRAM,
            label=ContactLabel.PRIMARY,
            value=normalize_telegram(
                self._cell(row, header_map, ImportFileHeader.TELEGRAM_PRIMARY)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.TELEGRAM,
            label=ContactLabel.SECONDARY,
            value=normalize_telegram(
                self._cell(row, header_map, ImportFileHeader.TELEGRAM_SECONDARY)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.VK,
            label="",
            value=normalize_handle(
                self._cell(row, header_map, ImportFileHeader.VK)
            ),
        )
        self._append_contact(
            contacts=contacts,
            contact_type=ContactType.MAX,
            label="",
            value=normalize_handle(
                self._cell(row, header_map, ImportFileHeader.MAX)
            ),
        )

        contacts = self._deduplicate_contacts(contacts)

        with transaction.atomic():
            client, created = Client.objects.update_or_create(
                external_id=external_id,
                defaults=client_defaults,
            )

            if created:
                stats.created += 1
            else:
                stats.updated += 1

            for contact in contacts:
                _, contact_created = ClientContact.objects.update_or_create(
                    client=client,
                    type=contact.type,
                    value=contact.value,
                    defaults={
                        "label": contact.label,
                        "status": ContactStatus.ACTIVE,
                        "metadata": contact.metadata,
                    },
                )

                if contact_created:
                    stats.contacts_created += 1
                else:
                    stats.contacts_updated += 1

    def _append_contact(
        self,
        *,
        contacts: list[ClientContactDTO],
        contact_type: str,
        label: str,
        value: str | None
    ) -> None:
        if not value:
            return

        contacts.append(
            ClientContactDTO(
                type=contact_type,
                label=label,
                value=value,
            )
        )

    def _deduplicate_contacts(self, contacts: list[ClientContactDTO]) -> list[ClientContactDTO]:
        unique: dict[tuple[str, str], ClientContactDTO] = {}

        for contact in contacts:
            key = (contact.type, contact.value)
            if key not in unique:
                unique[key] = contact

        return list(unique.values())

    def _parse_external_id(self, value: Any) -> int | None:
        if value in (None, ""):
            return None

        if isinstance(value, bool):
            raise ValueError("Некорректный № клиента")

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            if not value.is_integer():
                raise ValueError(f"Некорректный № клиента: {value}")
            return int(value)

        raw = clean_string(value)
        if not raw:
            return None

        try:
            return int(raw)
        except ValueError as exc:
            raise ValueError(f"Некорректный № клиента: {value}") from exc

    def _parse_date(self, value: Any) -> date | None:
        if value in (None, ""):
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        raw = clean_string(value)
        if not raw:
            return None

        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue

        raise ValueError(f"Не удалось распарсить дату: {value}")

    def _parse_bool(self, value: Any) -> bool:
        raw = (clean_string(value) or "").lower()
        return raw in {"1", "true", "yes", "y", "да", "+"}
